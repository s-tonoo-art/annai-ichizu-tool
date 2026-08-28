#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
案内図・位置図 自動作成ツール
================================

基地局の設計spec JSON（B_design_spec_*.json）と 表紙・目次・数量表 xlsm から
物件名・所在地・経度・緯度・標高・工事名・基地局番号・基地局名 を自動抽出し、
国土地理院タイル（標準地図）を用いて

    位置図  概ね 1:2500  （既定）
    案内図  概ね 1:25000 （既定）

の2枚の地図を生成し、A4横のPDF（案内図.pdf 相当のレイアウト）として出力する。

【使い方】
    python annai_ichizu_generator.py "<案件フォルダのパス>"

    案件フォルダ内に *.json（設計spec） と *.xlsm（数量表） が1つずつある前提で
    自動検出する。複数ある/見つからない場合は --json / --xlsm で明示指定する。

    例:
        python annai_ichizu_generator.py "raw\\2034_..._R004749901"
        python annai_ichizu_generator.py . --json spec.json --xlsm suryohyo.xlsm --out 案内図.pdf

【必要パッケージ】（事前に一度だけ）
    pip install reportlab requests pillow openpyxl

【地図データについて】
    国土地理院 地理院タイル（標準地図 std）を使用。出典表記はタイル画像内に
    "国土地理院" のクレジットが自動的に含まれる。インターネット接続が必要。
    https://maps.gsi.go.jp/development/ichiran.html

【数量表(xlsm)のセル対応】
    シート「入力シート」
        B2 = ①基地局番号
        B3 = ②局名（基地局名）
        B4 = ③工事件名（工事名）
    ここが空欄/未入力（"**" 等のプレースホルダ）の場合は --station-no / --station-name /
    --work-name で上書き指定できる。

【設計spec(json)のフィールド対応】
    responseBody.shared.site.billdingName.propertyNm      -> 物件名
    responseBody.shared.site.address.prefCd + cityCd       -> 所在地（都道府県+市区町村）
                                                               ※ jis_city_codes.csv で名称変換
    responseBody.shared.site.address.detailAddressNm       -> 所在地（丁目番地）
    responseBody.shared.site.degree.longitudeDeg           -> 東経（度→度分秒に変換）
    responseBody.shared.site.degree.latitudeDeg            -> 北緯（度→度分秒に変換）
    responseBody.shared.site.billding.groundLevel          -> 標高
"""

import argparse
import csv
import io
import json
import math
import sys
import time
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

try:
    import requests
except ImportError:
    print("requests が見つかりません。 pip install requests を実行してください。", file=sys.stderr)
    raise

try:
    from PIL import Image, ImageDraw
except ImportError:
    print("Pillow が見つかりません。 pip install pillow を実行してください。", file=sys.stderr)
    raise

try:
    import openpyxl
except ImportError:
    print("openpyxl が見つかりません。 pip install openpyxl を実行してください。", file=sys.stderr)
    raise

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
except ImportError:
    print("reportlab が見つかりません。 pip install reportlab を実行してください。", file=sys.stderr)
    raise

SCRIPT_DIR = Path(__file__).resolve().parent
CITY_CODE_CSV = SCRIPT_DIR / "jis_city_codes.csv"

GSI_TILE_URL = "https://cyberjapandata.gsi.go.jp/xyz/std/{z}/{x}/{y}.png"
TILE_SIZE = 256
GSI_MAX_ZOOM = 18
USER_AGENT = "annai-ichizu-generator/1.0 (internal tool; contact: user)"

MM_PER_PT = 25.4 / 72.0
PAGE_W_MM = 297.0
PAGE_H_MM = 210.0


# ----------------------------------------------------------------------
# 1. データ抽出
# ----------------------------------------------------------------------

@dataclass
class SiteData:
    property_nm: str = ""
    pref_nm: str = ""
    city_nm: str = ""
    detail_address: str = ""
    lat_deg: float = 0.0
    lon_deg: float = 0.0
    ground_level_m: float = 0.0

    work_name: str = ""
    station_no: str = ""
    station_name: str = ""

    @property
    def address_full(self) -> str:
        return f"{self.pref_nm}　{self.city_nm}　{self.detail_address}".strip("　 ")


def load_city_codes(csv_path: Path) -> dict:
    table = {}
    if not csv_path.exists():
        return table
    with open(csv_path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            table[row["city_cd5"]] = (row["pref_nm"], row["city_nm"])
    return table


def extract_from_json(json_path: Path) -> SiteData:
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    site = data["responseBody"]["shared"]["site"]

    sd = SiteData()
    sd.property_nm = site.get("billdingName", {}).get("propertyNm", "") or ""
    sd.detail_address = site.get("address", {}).get("detailAddressNm", "") or ""
    sd.lat_deg = float(site["degree"]["latitudeDeg"])
    sd.lon_deg = float(site["degree"]["longitudeDeg"])
    sd.ground_level_m = float(site.get("billding", {}).get("groundLevel", 0.0) or 0.0)

    pref_cd = str(site.get("address", {}).get("prefCd", "")).zfill(2)
    city_cd_local = str(site.get("address", {}).get("cityCd", "")).zfill(3)
    full_code = f"{pref_cd}{city_cd_local}"

    city_table = load_city_codes(CITY_CODE_CSV)
    if full_code in city_table:
        sd.pref_nm, sd.city_nm = city_table[full_code]
    else:
        sd.pref_nm, sd.city_nm = f"(pref:{pref_cd})", f"(city:{city_cd_local})"
        print(f"[警告] 市区町村コード {full_code} が jis_city_codes.csv に見つかりません。"
              f" pref_cd/city_cd を確認してください。", file=sys.stderr)

    return sd


def extract_from_xlsm(xlsm_path: Path, sd: SiteData) -> None:
    """入力シート B2/B3/B4 から 基地局番号・局名・工事件名 を取得して sd に反映する。"""
    wb = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=False)
    if "入力シート" not in wb.sheetnames:
        print("[警告] xlsm に「入力シート」が見つかりません。--station-no 等で指定してください。",
              file=sys.stderr)
        return
    ws = wb["入力シート"]
    station_no = ws["B2"].value
    station_name = ws["B3"].value
    work_name = ws["B4"].value

    def clean(v):
        if v is None:
            return ""
        v = str(v).strip()
        return "" if v in ("**", "***", "-") else v

    sd.station_no = clean(station_no)
    sd.station_name = clean(station_name)
    sd.work_name = clean(work_name)


def deg_to_dms(deg: float) -> str:
    """10進度 -> '135度29分13.47秒' 形式"""
    d = int(math.floor(deg))
    m_full = (deg - d) * 60
    m = int(math.floor(m_full))
    s = (m_full - m) * 60
    return f"{d}度{m}分{s:05.2f}秒"


# ----------------------------------------------------------------------
# 2. 地理院タイル取得・地図画像作成
# ----------------------------------------------------------------------

def latlon_to_world_px(lat: float, lon: float, zoom: int) -> tuple:
    lat_rad = math.radians(lat)
    n = 2 ** zoom
    x = (lon + 180.0) / 360.0 * n * TILE_SIZE
    y = (1.0 - math.log(math.tan(lat_rad) + 1.0 / math.cos(lat_rad)) / math.pi) / 2.0 * n * TILE_SIZE
    return x, y


def meters_per_pixel(lat: float, zoom: int) -> float:
    return 156543.03392804097 * math.cos(math.radians(lat)) / (2 ** zoom)


def choose_zoom(lat: float, ground_width_m: float, target_px: float, max_zoom: int = GSI_MAX_ZOOM) -> int:
    """目標px数に対して、解像度が足りる（=誤差の少ない）"最小"のズームを選ぶ
    （ズームが上がるほどmeters_per_pixelは単調減少するので、要求解像度を満たす
    最小のzを小さい方から探す。大きい方から探すと常にmax_zoomが返ってしまうバグに注意）"""
    required_mpp = ground_width_m / target_px
    for z in range(1, max_zoom + 1):
        if meters_per_pixel(lat, z) <= required_mpp:
            return z
    return max_zoom


_network_dead = False  # 最初の通信失敗以降はリトライをスキップして高速に諦める


def fetch_tile(session: requests.Session, z: int, x: int, y: int, cache_dir: Path,
                offline: bool = False) -> Image.Image:
    global _network_dead
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_path = cache_dir / f"{z}_{x}_{y}.png"
    if cache_path.exists():
        return Image.open(cache_path).convert("RGB")

    n = 2 ** z
    if x < 0 or x >= n or y < 0 or y >= n:
        # 地図範囲外 -> 白タイルで代用（キャッシュはしない）
        return Image.new("RGB", (TILE_SIZE, TILE_SIZE), "white")

    if offline or _network_dead:
        return Image.new("RGB", (TILE_SIZE, TILE_SIZE), (230, 230, 220))

    url = GSI_TILE_URL.format(z=z, x=x, y=y)
    attempts = 2
    connection_failed = False
    for attempt in range(attempts):
        try:
            resp = session.get(url, timeout=6, headers={"User-Agent": USER_AGENT})
            if resp.status_code == 200:
                img = Image.open(io.BytesIO(resp.content)).convert("RGB")
                img.save(cache_path)
                return img
        except requests.RequestException:
            connection_failed = True
        if attempt < attempts - 1:
            time.sleep(0.3)
    print(f"[警告] タイル取得失敗: z={z} x={x} y={y}", file=sys.stderr)
    if connection_failed:
        # 接続自体ができない = 以降も同じなので即座に諦めて高速化する
        _network_dead = True
    _network_dead_hint()
    return Image.new("RGB", (TILE_SIZE, TILE_SIZE), (230, 230, 220))


_hint_shown = False


def _network_dead_hint():
    """最初の失敗タイルで、以降のタイルもまとめて諦めるかの目安を出す（連続失敗時のみ）"""
    global _hint_shown
    if not _hint_shown:
        _hint_shown = True
        print("[警告] 地理院タイルサーバーに到達できません。社内プロキシ/ファイアウォールで "
              "cyberjapandata.gsi.go.jp への通信がブロックされていないか確認してください。"
              "（このまま続行し、以降のタイルは灰色のプレースホルダーで埋めます）", file=sys.stderr)


def build_map_image(lat: float, lon: float, ground_width_m: float, ground_height_m: float,
                     out_px_w: int, out_px_h: int, cache_dir: Path,
                     session: requests.Session, offline: bool = False) -> Image.Image:
    """指定した中心座標・実距離(m)・出力ピクセルサイズで地理院タイルから地図画像を切り出す"""
    zoom = choose_zoom(lat, max(ground_width_m, ground_height_m), max(out_px_w, out_px_h))

    center_x, center_y = latlon_to_world_px(lat, lon, zoom)
    mpp = meters_per_pixel(lat, zoom)
    half_w_px = (ground_width_m / mpp) / 2.0
    half_h_px = (ground_height_m / mpp) / 2.0

    left = center_x - half_w_px
    top = center_y - half_h_px
    right = center_x + half_w_px
    bottom = center_y + half_h_px

    tile_x0 = int(math.floor(left / TILE_SIZE))
    tile_x1 = int(math.floor(right / TILE_SIZE))
    tile_y0 = int(math.floor(top / TILE_SIZE))
    tile_y1 = int(math.floor(bottom / TILE_SIZE))

    canvas_w = (tile_x1 - tile_x0 + 1) * TILE_SIZE
    canvas_h = (tile_y1 - tile_y0 + 1) * TILE_SIZE
    big = Image.new("RGB", (canvas_w, canvas_h), "white")

    for tx in range(tile_x0, tile_x1 + 1):
        for ty in range(tile_y0, tile_y1 + 1):
            tile = fetch_tile(session, zoom, tx, ty, cache_dir, offline=offline)
            big.paste(tile, ((tx - tile_x0) * TILE_SIZE, (ty - tile_y0) * TILE_SIZE))

    crop_left = left - tile_x0 * TILE_SIZE
    crop_top = top - tile_y0 * TILE_SIZE
    crop_box = (
        int(round(crop_left)),
        int(round(crop_top)),
        int(round(crop_left + (right - left))),
        int(round(crop_top + (bottom - top))),
    )
    cropped = big.crop(crop_box)
    resized = cropped.resize((out_px_w, out_px_h), Image.LANCZOS)
    return resized


def draw_crosshair_marker(img: Image.Image) -> Image.Image:
    """画像中心に赤丸+黒十字のマーカーを描画（案内図.pdf のマーカーを模したもの）"""
    img = img.copy()
    draw = ImageDraw.Draw(img)
    w, h = img.size
    cx, cy = w / 2.0, h / 2.0
    r = min(w, h) * 0.045
    cross_ext = r * 1.55

    draw.ellipse((cx - r, cy - r, cx + r, cy + r), outline=(230, 20, 20), width=max(2, int(w * 0.004)))
    lw = max(2, int(w * 0.0035))
    draw.line((cx - cross_ext, cy, cx + cross_ext, cy), fill=(0, 0, 0), width=lw)
    draw.line((cx, cy - cross_ext, cx, cy + cross_ext), fill=(0, 0, 0), width=lw)
    return img


# ----------------------------------------------------------------------
# 3. PDF レイアウト（案内図.pdf のレイアウトを踏襲）
# ----------------------------------------------------------------------

# 案内図.pdf を実測して得た座標（pt, 左上原点）。この座標系のまま扱い、
# 描画時にreportlab座標（左下原点）へ変換する。
LAYOUT = {
    "outer_frame": (38.9, 20.2, 806.5, 568.7),
    "map_ichizu": (69.0, 63.4, 372.0, 366.4),      # 位置図
    "map_annai": (449.3, 63.4, 752.4, 366.4),      # 案内図
    "label_ichizu": (190.0, 375.7, 251.0, 412.4),
    "label_annai": (565.7, 375.7, 635.9, 412.4),
    "attr_table": (308.0, 409.0, 519.4, 463.7),    # 物件名/所在地/東経/北緯/標高
    "attr_table_col_split": 344.6,
    "title_block": (393.1, 520.6, 806.5, 568.7),
    "handling_box": (682.9, 30.6, 796.4, 55.7),
}


def pt2mm(v: float) -> float:
    return v * MM_PER_PT


def topdown_to_rl_y(y_topdown_pt: float) -> float:
    """PDF左上原点(pt, 下向き+) の y座標を reportlab(左下原点, 上向き+, mm) の y座標に変換"""
    return PAGE_H_MM - pt2mm(y_topdown_pt)


def register_japanese_font():
    """日本語フォントを登録する。

    優先順位:
      1) Windowsの実フォント（メイリオ/游ゴシック/MSゴシック）が見つかればそれを埋め込む
         （TrueTypeのみ対応。OpenType/CFF系フォントはreportlabのTTFontで読めないため不可）
      2) reportlab標準搭載の日本語CIDフォント(HeiseiKakuGo-W5)を使う
         -> 追加ファイル不要でどの環境でも必ず表示できる（Acrobat等のCJKフォント代替で描画）
    """
    font_candidates = [
        r"C:\Windows\Fonts\meiryo.ttc",
        r"C:\Windows\Fonts\YuGothM.ttc",
        r"C:\Windows\Fonts\msgothic.ttc",
    ]
    for path in font_candidates:
        p = Path(path)
        if p.exists():
            try:
                pdfmetrics.registerFont(TTFont("JP", str(p), subfontIndex=0))
                return "JP"
            except Exception:
                continue

    try:
        pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
        return "HeiseiKakuGo-W5"
    except Exception:
        pass

    print("[警告] 日本語フォントの登録に失敗しました。文字が正しく表示されない可能性があります。",
          file=sys.stderr)
    return "Helvetica"


def compose_pdf(sd: SiteData, ichizu_img: Image.Image, annai_img: Image.Image,
                 out_path: Path, work_no: str, sheet_no: str = "01",
                 scale_ichizu: int = 2500, scale_annai: int = 25000,
                 rev_date: str = None, tmp_dir: Path = None):
    tmp_dir = tmp_dir or out_path.parent
    ichizu_png = tmp_dir / "_tmp_ichizu.png"
    annai_png = tmp_dir / "_tmp_annai.png"
    ichizu_img.save(ichizu_png)
    annai_img.save(annai_png)

    font_name = register_japanese_font()
    rev_date = rev_date or date.today().strftime("%Y/%m/%d")

    c = canvas.Canvas(str(out_path), pagesize=landscape(A4))
    c.setFont(font_name, 9)

    def rect_td(box_pt, stroke=1, fill=None):
        x0, y0, x1, y1 = box_pt
        rl_y0 = topdown_to_rl_y(y0)  # top edge -> higher rl_y
        rl_y1 = topdown_to_rl_y(y1)  # bottom edge -> lower rl_y
        c.saveState()
        if fill:
            c.setFillColorRGB(*fill)
            c.rect(pt2mm(x0) * mm, rl_y1 * mm, pt2mm(x1 - x0) * mm, (rl_y0 - rl_y1) * mm,
                   stroke=stroke, fill=1)
        else:
            c.rect(pt2mm(x0) * mm, rl_y1 * mm, pt2mm(x1 - x0) * mm, (rl_y0 - rl_y1) * mm,
                   stroke=stroke, fill=0)
        c.restoreState()

    def line_td(x0, y0, x1, y1):
        c.line(pt2mm(x0) * mm, topdown_to_rl_y(y0) * mm, pt2mm(x1) * mm, topdown_to_rl_y(y1) * mm)

    def text_center_td(cx_pt, cy_pt, s, size=10):
        c.setFont(font_name, size)
        c.drawCentredString(pt2mm(cx_pt) * mm, topdown_to_rl_y(cy_pt) * mm, s)

    def text_left_td(x_pt, y_pt, s, size=8):
        c.setFont(font_name, size)
        c.drawString(pt2mm(x_pt) * mm, topdown_to_rl_y(y_pt) * mm, s)

    from reportlab.lib.units import mm  # local import so `mm` symbol is defined for above closures

    # 外枠
    rect_td(LAYOUT["outer_frame"])

    # 位置図 / 案内図 画像
    ix0, iy0, ix1, iy1 = LAYOUT["map_ichizu"]
    c.drawImage(str(ichizu_png), pt2mm(ix0) * mm, topdown_to_rl_y(iy1) * mm,
                pt2mm(ix1 - ix0) * mm, pt2mm(iy1 - iy0) * mm)
    rect_td(LAYOUT["map_ichizu"])

    ax0, ay0, ax1, ay1 = LAYOUT["map_annai"]
    c.drawImage(str(annai_png), pt2mm(ax0) * mm, topdown_to_rl_y(ay1) * mm,
                pt2mm(ax1 - ax0) * mm, pt2mm(ay1 - ay0) * mm)
    rect_td(LAYOUT["map_annai"])

    # ラベル
    lx0, ly0, lx1, ly1 = LAYOUT["label_ichizu"]
    cx = (lx0 + lx1) / 2
    text_center_td(cx, ly0 + 14, "位置図", size=13)
    text_center_td(cx, ly0 + 30, f"1：{scale_ichizu:,}", size=13)

    lx0, ly0, lx1, ly1 = LAYOUT["label_annai"]
    cx = (lx0 + lx1) / 2
    text_center_td(cx, ly0 + 14, "案内図", size=13)
    text_center_td(cx, ly0 + 30, f"1：{scale_annai:,}", size=13)

    # 属性テーブル（物件名/所在地/東経/北緯/標高）
    tx0, ty0, tx1, ty1 = LAYOUT["attr_table"]
    col_split = LAYOUT["attr_table_col_split"]
    n_rows = 5
    row_h = (ty1 - ty0) / n_rows
    labels = ["物件名", "所在地", "東経", "北緯", "標高"]
    values = [
        sd.property_nm,
        sd.address_full,
        deg_to_dms(sd.lon_deg),
        deg_to_dms(sd.lat_deg),
        f"{sd.ground_level_m:+.0f}m",
    ]
    rect_td(LAYOUT["attr_table"])
    line_td(col_split, ty0, col_split, ty1)
    for i in range(1, n_rows):
        y = ty0 + row_h * i
        line_td(tx0, y, tx1, y)
    for i, (lab, val) in enumerate(zip(labels, values)):
        row_top = ty0 + row_h * i
        row_mid = row_top + row_h / 2
        text_center_td((tx0 + col_split) / 2, row_mid + 3, lab, size=8)
        text_left_td(col_split + 4, row_mid + 3, val, size=8)

    # 表題欄（右下）: 案内図.pdf を実測して求めた6段グリッド（等間隔）を基準に配置する
    bx0, by0, bx1, by1 = LAYOUT["title_block"]
    rect_td(LAYOUT["title_block"])
    unit = (by1 - by0) / 6.0
    L = [by0 + unit * i for i in range(7)]  # L[0]..L[6]

    # ---- 左側: 版数・日付・改定履歴 の小表（3列 x 6段：見出し+5行） ----
    col_l1, col_l2 = 408.2, 438.4
    line_td(col_l1, by0, col_l1, by1)
    line_td(col_l2, by0, col_l2, by1)
    for y in L[1:-1]:
        line_td(bx0, y, 544.6, y)
    text_center_td((bx0 + col_l1) / 2, L[0] + unit / 2 + 3, "版数", size=6)
    text_center_td((col_l1 + col_l2) / 2, L[0] + unit / 2 + 3, "日付", size=6)
    text_center_td((col_l2 + 544.6) / 2, L[0] + unit / 2 + 3, "改定履歴", size=6)
    text_center_td((bx0 + col_l1) / 2, L[1] + unit / 2 + 3, "1", size=7)
    text_center_td((col_l1 + col_l2) / 2, L[1] + unit / 2 + 2.5, rev_date, size=4.6)
    text_center_td((col_l2 + 544.6) / 2, L[1] + unit / 2 + 3, "初版", size=7)
    for i in range(2, 6):
        text_center_td((bx0 + col_l1) / 2, L[i] + unit / 2 + 3, str(i), size=7)

    # ---- 右側 ----
    col2 = 544.6   # 左小表との境界
    col3 = 571.1   # 日付値/縮尺値 の開始列
    col4 = 626.2   # 工事名列・基地局番号列・図面名列 の開始
    col5 = 661.2   # 工事名値・基地局名値・図面名値 の開始
    col6 = 760.0   # 図面番号ラベルの開始
    col7 = 793.4   # 図面番号 値(01) の開始

    line_td(col2, by0, col2, by1)
    line_td(col3, L[0], col3, L[2])
    line_td(col4, by0, col4, by1)
    line_td(col5, L[0], col5, L[4])
    line_td(col6, L[4], col6, by1)
    line_td(col7, L[4], col7, by1)

    # 横の区切り
    line_td(col2, L[2], bx1, L[2])          # R0/R1 境界（全体）
    line_td(col2, L[3], col4, L[3])         # R1/R2 境界（縮尺/施主 列のみ）
    line_td(col2, L[4], col4, L[4])         # R2/R3 境界（施主/ソフトバンク列のみ）
    line_td(col4, L[4], bx1, L[4])          # R2/R3 境界（図面名/図面番号 列）

    # R0: 日付 / 工事名（2段ぶんの高さ）
    r0_mid = L[0] + unit
    text_center_td((col2 + col3) / 2, r0_mid + 3, "日付", size=6.5)
    text_center_td((col3 + col4) / 2, r0_mid + 3, rev_date, size=7)
    text_center_td((col4 + col5) / 2, r0_mid + 3, "工事名", size=6.5)
    text_left_td(col5 + 3, r0_mid + 3, sd.work_name or "-", size=7.5)

    # R1: 縮尺 / 基地局番号・基地局名（見出しは2段結合）
    r1_mid = L[2] + unit / 2
    text_center_td((col2 + col3) / 2, r1_mid + 3, "縮尺", size=6.5)
    text_center_td((col3 + col4) / 2, r1_mid + 3, "図示", size=7)

    # R2: 施主(見出しのみ)
    r2_mid = L[3] + unit / 2
    text_center_td((col2 + col4) / 2, r2_mid + 3, "施主", size=6.5)

    # 基地局番号・基地局名（R1+R2結合セル）
    station_mid = L[2] + unit
    text_center_td((col4 + col5) / 2, L[2] + unit / 2 + 3, "基地局番号", size=6)
    text_center_td((col4 + col5) / 2, L[3] + unit / 2 + 3, "基地局名", size=6)
    text_left_td(col5 + 3, station_mid + 3, f"{sd.station_no or '-'}　{sd.station_name or '-'}", size=7.5)

    # R3: ソフトバンク株式会社（施主値, 2段ぶん結合） / 図面名 / 図面番号
    r3_mid = L[4] + unit
    text_center_td((col2 + col4) / 2, r3_mid + 3, "ソフトバンク株式会社", size=7.5)
    text_center_td((col4 + col5) / 2, r3_mid + 3, "図面名", size=6.5)
    text_left_td(col5 + 3, r3_mid + 3, "案内図・位置図", size=7.5)
    text_center_td((col5 + col6) / 2, r3_mid + 3, "", size=6.5)
    text_center_td((col6 + col7) / 2, r3_mid + 3, "図面番号", size=6.5)
    text_center_td((col7 + bx1) / 2, r3_mid + 3, sheet_no, size=8)

    # 取扱注意
    rect_td(LAYOUT["handling_box"])
    hx0, hy0, hx1, hy1 = LAYOUT["handling_box"]
    text_center_td((hx0 + hx1) / 2, (hy0 + hy1) / 2 + 5, "取 扱 注 意", size=11)

    c.showPage()
    c.save()

    ichizu_png.unlink(missing_ok=True)
    annai_png.unlink(missing_ok=True)


# ----------------------------------------------------------------------
# 4. メイン
# ----------------------------------------------------------------------

def autodetect(folder: Path, pattern: str, label: str) -> Path:
    matches = sorted(folder.glob(pattern))
    if not matches:
        raise FileNotFoundError(f"{label} が見つかりません（{folder} 内に {pattern}）。--json/--xlsm で指定してください。")
    if len(matches) > 1:
        print(f"[警告] {label} が複数見つかりました。先頭を使用します: {matches[0].name}", file=sys.stderr)
    return matches[0]


def main():
    ap = argparse.ArgumentParser(description="案内図・位置図 自動作成ツール")
    ap.add_argument("folder", type=str, help="設計spec(json)と数量表(xlsm)が入っている案件フォルダ")
    ap.add_argument("--json", type=str, default=None, help="設計spec jsonファイルを明示指定")
    ap.add_argument("--xlsm", type=str, default=None, help="数量表 xlsmファイルを明示指定")
    ap.add_argument("--out", type=str, default=None, help="出力PDFパス（既定: <folder>/案内図.pdf）")
    ap.add_argument("--scale-ichizu", type=int, default=2500, help="位置図の縮尺分母（既定2500）")
    ap.add_argument("--scale-annai", type=int, default=25000, help="案内図の縮尺分母（既定25000）")
    ap.add_argument("--dpi", type=int, default=200, help="地図画像の出力解像度目安（既定200dpi）")
    ap.add_argument("--sheet-no", type=str, default="01", help="図面番号（既定 01）")
    ap.add_argument("--station-no", type=str, default=None, help="基地局番号を上書き指定")
    ap.add_argument("--station-name", type=str, default=None, help="基地局名を上書き指定")
    ap.add_argument("--work-name", type=str, default=None, help="工事名を上書き指定")
    ap.add_argument("--cache-dir", type=str, default=None, help="タイルキャッシュ保存先（既定: <folder>/.tile_cache）")
    ap.add_argument("--offline", action="store_true",
                     help="地理院タイルを取得せず灰色プレースホルダーでレイアウトのみ確認する")
    args = ap.parse_args()

    folder = Path(args.folder).resolve()
    if not folder.exists():
        print(f"フォルダが見つかりません: {folder}", file=sys.stderr)
        sys.exit(1)

    json_path = Path(args.json) if args.json else autodetect(folder, "*.json", "設計spec(json)")
    xlsm_path = Path(args.xlsm) if args.xlsm else autodetect(folder, "*.xlsm", "数量表(xlsm)")
    out_path = Path(args.out) if args.out else folder / "案内図.pdf"
    cache_dir = Path(args.cache_dir) if args.cache_dir else folder / ".tile_cache"

    print(f"[INFO] json  : {json_path}")
    print(f"[INFO] xlsm  : {xlsm_path}")
    print(f"[INFO] out   : {out_path}")

    sd = extract_from_json(json_path)
    extract_from_xlsm(xlsm_path, sd)

    if args.station_no:
        sd.station_no = args.station_no
    if args.station_name:
        sd.station_name = args.station_name
    if args.work_name:
        sd.work_name = args.work_name

    print(f"[INFO] 物件名   : {sd.property_nm}")
    print(f"[INFO] 所在地   : {sd.address_full}")
    print(f"[INFO] 緯度/経度: {sd.lat_deg} / {sd.lon_deg}")
    print(f"[INFO] 標高     : {sd.ground_level_m}")
    print(f"[INFO] 基地局番号: {sd.station_no or '(未取得)'}")
    print(f"[INFO] 基地局名 : {sd.station_name or '(未取得)'}")
    print(f"[INFO] 工事名   : {sd.work_name or '(未取得)'}")

    # 地図ボックスの実寸(mm) -> 地上距離(m)
    box_ichizu = LAYOUT["map_ichizu"]
    box_annai = LAYOUT["map_annai"]
    box_w_mm_ichizu = pt2mm(box_ichizu[2] - box_ichizu[0])
    box_h_mm_ichizu = pt2mm(box_ichizu[3] - box_ichizu[1])
    box_w_mm_annai = pt2mm(box_annai[2] - box_annai[0])
    box_h_mm_annai = pt2mm(box_annai[3] - box_annai[1])

    ground_w_ichizu = box_w_mm_ichizu / 1000.0 * args.scale_ichizu
    ground_h_ichizu = box_h_mm_ichizu / 1000.0 * args.scale_ichizu
    ground_w_annai = box_w_mm_annai / 1000.0 * args.scale_annai
    ground_h_annai = box_h_mm_annai / 1000.0 * args.scale_annai

    px_w_ichizu = round(box_w_mm_ichizu / 25.4 * args.dpi)
    px_h_ichizu = round(box_h_mm_ichizu / 25.4 * args.dpi)
    px_w_annai = round(box_w_mm_annai / 25.4 * args.dpi)
    px_h_annai = round(box_h_mm_annai / 25.4 * args.dpi)

    print(f"[INFO] 位置図 地上範囲: {ground_w_ichizu:.1f}m x {ground_h_ichizu:.1f}m -> {px_w_ichizu}x{px_h_ichizu}px")
    print(f"[INFO] 案内図 地上範囲: {ground_w_annai:.1f}m x {ground_h_annai:.1f}m -> {px_w_annai}x{px_h_annai}px")

    session = requests.Session()

    print("[INFO] 位置図タイル取得中..." if not args.offline else "[INFO] 位置図（オフラインプレビュー）作成中...")
    ichizu_img = build_map_image(sd.lat_deg, sd.lon_deg, ground_w_ichizu, ground_h_ichizu,
                                  px_w_ichizu, px_h_ichizu, cache_dir, session, offline=args.offline)
    ichizu_img = draw_crosshair_marker(ichizu_img)

    print("[INFO] 案内図タイル取得中..." if not args.offline else "[INFO] 案内図（オフラインプレビュー）作成中...")
    annai_img = build_map_image(sd.lat_deg, sd.lon_deg, ground_w_annai, ground_h_annai,
                                 px_w_annai, px_h_annai, cache_dir, session, offline=args.offline)
    annai_img = draw_crosshair_marker(annai_img)

    print("[INFO] PDF作成中...")
    compose_pdf(sd, ichizu_img, annai_img, out_path,
                work_no=sd.work_name, sheet_no=args.sheet_no,
                scale_ichizu=args.scale_ichizu, scale_annai=args.scale_annai,
                tmp_dir=folder)

    print(f"[完了] {out_path}")


if __name__ == "__main__":
    main()
